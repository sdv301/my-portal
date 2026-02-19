# reports/template_report_generator.py
import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime, date, timedelta
import json
import re
from copy import copy

class TemplateReportGenerator:
    def __init__(self, db_connection, template_path: str = None):
        self.db = db_connection
        self.reports_dir = 'reports_output'
        os.makedirs(self.reports_dir, exist_ok=True)
        
        if template_path is None:
            self.template_path = 'report_templates/Сводный_отчет_шаблон.xlsx'
        else:
            self.template_path = template_path
        
        if not os.path.exists(self.template_path):
            possible_paths = [
                'report_templates/Сводный_отчет_шаблон.xlsx',
                '../report_templates/Сводный_отчет_шаблон.xlsx',
                './report_templates/Сводный_отчет_шаблон.xlsx'
            ]
            for path in possible_paths:
                if os.path.exists(path):
                    self.template_path = path
                    break
            else:
                raise FileNotFoundError(f"Шаблон не найден. Искал: {possible_paths}")

    def generate_report(self, report_date: date = None) -> str:
        try:
            if report_date is None:
                report_date = datetime.now().date()

            print(f"\n🎯 ГЕНЕРАЦИЯ ОТЧЕТА НА {report_date.strftime('%d.%m.%Y')}")

            aggregated_data = self.db.get_aggregated_data()
            if not aggregated_data:
                raise Exception("Нет данных в БД")

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'Сводный_отчет_{timestamp}.xlsx'
            output_path = os.path.join(self.reports_dir, filename)
            
            shutil.copy2(self.template_path, output_path)

            wb = load_workbook(output_path)
            self._update_report_info(wb, report_date)
            self._fill_all_company_data(wb, aggregated_data, report_date)
            wb.save(output_path)

            if os.path.exists(output_path):
                print(f"✅ Отчет создан успешно: {output_path}")
                return output_path
            else:
                raise Exception("Файл не был создан")
        except Exception as e:
            print(f"❌ Ошибка: {e}")
            raise

    def _copy_style(self, source_cell, target_cell):
        """Копирует форматирование из эталонной ячейки и включает перенос текста"""
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            
            # Копируем выравнивание, но жестко задаем wrap_text=True, чтобы текст не вылезал из ячейки
            al = source_cell.alignment
            if al:
                target_cell.alignment = Alignment(
                    horizontal=al.horizontal,
                    vertical=al.vertical,
                    text_rotation=al.text_rotation,
                    wrap_text=True, 
                    shrink_to_fit=al.shrink_to_fit,
                    indent=al.indent
                )

    def _set_cell_value(self, ws, row: int, col: int, value, template_row: int = None):
        """Устанавливает значение ячейки и копирует стиль, если указана строка-шаблон"""
        try:
            if value is None: 
                value = 0 if isinstance(value, (int, float)) else ""
            if row > 0 and col > 0:
                target_cell = ws.cell(row=row, column=col)
                target_cell.value = value
                
                # Если передан номер строки шаблона, копируем из неё стиль
                if template_row and row != template_row:
                    source_cell = ws.cell(row=template_row, column=col)
                    self._copy_style(source_cell, target_cell)
                else:
                    # Даже если нет шаблона, предотвращаем вылезание текста
                    al = target_cell.alignment
                    if al:
                        target_cell.alignment = Alignment(
                            horizontal=al.horizontal,
                            vertical=al.vertical,
                            text_rotation=al.text_rotation,
                            wrap_text=True,
                            shrink_to_fit=al.shrink_to_fit,
                            indent=al.indent
                        )
                    else:
                        target_cell.alignment = Alignment(wrap_text=True)
                return True
        except: 
            return False

    def _get_supplier_string(self, company_name: str) -> str:
        """Определяет строку поставщиков на основе имени компании"""
        name_lower = str(company_name).lower()
        if 'саханефтегазсбыт' in name_lower or 'снгс' in name_lower:
            return 'ООО Газпромнефть-РП (Омская НПЗ)'
        elif 'туймаада' in name_lower:
            return 'Ангарский НПЗ, Ачинский НПЗ, Омский НПЗ, Сургутский ЗСК '
        elif 'сибойл' in name_lower:
            return 'ПАО "НК "Роснефть",ООО "Татнефть-АЗС Центр",ООО "ГАЗПРОМ ГАЗОНЕФТЕПРОДУКТ ПРОДАЖИ",ПАО "Газпром нефть"'
        elif 'экто' in name_lower:
            return 'ПАО "НК "Роснефть",ООО "Татнефть-АЗС Центр",ООО "ГАЗПРОМ ГАЗОНЕФТЕПРОДУКТ ПРОДАЖИ",ПАО "Газпром нефть"'
        elif 'паритет' in name_lower:
            return 'Стандарт, ТЭК Восток, Синергия, ПетроТрейд, ПетроТрейд, Миком'
        return ''

    def _get_oil_depot_string(self, company_name: str) -> str:
        """Определяет строку нефтебаз (Колонка C) для Листа 4 на основе имени компании"""
        name_lower = str(company_name).lower()
        if 'саханефтегазсбыт' in name_lower or 'снгс' in name_lower:
            return 'НБ Батагайская, НБ Белогорская, НБ Жиганская,НБ Зырянская,НБ Ленская, НБ Нагорнинская, НБ Нижне-Бестяхская, НБ Нижнеколымская,НБ Нижнеянская, НБ Нюрбинская,НБ Олекминская, НБ Сангарская, НБ Среднеколымская, НБ Томмотская, НБ Усть- Куйгинская, НБ Хандыгская, НБ Чокурдахская, НБ Эльдиканская, НБ Якутская       '
        elif 'туймаада' in name_lower:
            return 'Нижне-Бестяхская нефтебаза, Сунтарская нефтебаза (договор хранения), Нюрбинская нефтебаза (договор хранения), Якутская нефтебаза (договор хранения)'
        elif 'сибойл' in name_lower:
            return 'АО НК "Туймаада-Нефть"'
        elif 'паритет' in name_lower:
            return 'ООО "Дорснаб", ООО "Экресурс"'
        return 'Все объекты'

    def _update_report_info(self, wb, report_date: date):
        date_str = report_date.strftime('%d.%m.%Y')
        past_date_str = (report_date - timedelta(days=1)).strftime('%d.%m.%Y')
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in range(1, 6):
                for col in range(1, 10):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and 'дата' in str(cell.value).lower():
                        ws.cell(row=row, column=col+1).value = date_str

        # Особое правило для Листа 1 - дата минус 1 день
        if '1-Структура' in wb.sheetnames:
            ws1 = wb['1-Структура']
            found = False
            for r in range(1, 5):
                cell = ws1.cell(row=r, column=1)
                if cell.value and 'состоянию' in str(cell.value).lower():
                    ws1.cell(row=r, column=2).value = past_date_str
                    found = True
                    break
            
            if not found:
                ws1.cell(row=1, column=2).value = past_date_str

    def _fill_all_company_data(self, wb, aggregated_data: dict, report_date: date):
        if '1-Структура' in wb.sheetnames:
            self._fill_structure_sheet_full(wb['1-Структура'], aggregated_data)
        if '2-Потребность' in wb.sheetnames:
            self._fill_demand_sheet_full(wb['2-Потребность'], aggregated_data)
        if '3-Остатки' in wb.sheetnames:
            self._fill_stocks_sheet_full(wb['3-Остатки'], aggregated_data)
        if '4-Поставка' in wb.sheetnames:
            self._fill_supply_sheet_full(wb['4-Поставка'], aggregated_data, report_date)
        if '5-Реализация' in wb.sheetnames:
            self._fill_sales_sheet_full(wb['5-Реализация'], aggregated_data)
        if '6-Авиатопливо' in wb.sheetnames:
            self._fill_aviation_sheet_full(wb['6-Авиатопливо'], aggregated_data)
        
        # Лист 7 оставлен закомментированным, чтобы он не перезаписывался и оставался статичным из шаблона
        # if '7-Комментарии' in wb.sheetnames:
        #     self._fill_comments_sheet_full(wb['7-Комментарии'], aggregated_data)

    def _fill_structure_sheet_full(self, ws, aggregated_data: dict):
        start_row = 13
        current_row = start_row
        for company_name, company_data in aggregated_data.items():
            for record in company_data.get('sheet1', []):
                if 'наименование компаний' in str(record.get('company_name', '')).lower(): continue
                self._set_cell_value(ws, current_row, 1, record.get('affiliation', ''), start_row)
                self._set_cell_value(ws, current_row, 2, record.get('company_name', company_name), start_row)
                self._set_cell_value(ws, current_row, 3, record.get('oil_depots_count', 0), start_row)
                self._set_cell_value(ws, current_row, 4, record.get('azs_count', 0), start_row)
                self._set_cell_value(ws, current_row, 5, record.get('working_azs_count', 0), start_row)
                current_row += 1

    def _fill_demand_sheet_full(self, ws, aggregated_data: dict):
        year_row = 7
        month_row = 13
        cur_year_row = year_row
        cur_month_row = month_row
        for company_name, company_data in aggregated_data.items():
            data = company_data.get('sheet2', {})
            if data:
                self._set_cell_value(ws, cur_year_row, 1, company_name, year_row)
                self._set_cell_value(ws, cur_year_row, 4, data.get('gasoline_ai92', 0), year_row)
                self._set_cell_value(ws, cur_year_row, 5, data.get('gasoline_ai95', 0), year_row)
                self._set_cell_value(ws, cur_year_row, 8, data.get('diesel_total', 0), year_row)
                
                self._set_cell_value(ws, cur_month_row, 1, company_name, month_row)
                self._set_cell_value(ws, cur_month_row, 4, data.get('monthly_gasoline_total', 0) / 2 if data.get('monthly_gasoline_total') else 0, month_row)
                self._set_cell_value(ws, cur_month_row, 5, data.get('monthly_gasoline_total', 0) / 2 if data.get('monthly_gasoline_total') else 0, month_row)
                self._set_cell_value(ws, cur_month_row, 8, data.get('monthly_diesel_total', 0), month_row)
                cur_year_row += 1
                cur_month_row += 1

    def _fill_stocks_sheet_full(self, ws, aggregated_data: dict):
        start_row = 9
        current_row = start_row
        for company_name, company_data in aggregated_data.items():
            sheet3_recs = company_data.get('sheet3_data', [])
            
            azs_count = 0
            azs_totals = {
                'stock_ai92': 0, 'stock_ai95': 0, 'stock_ai98_ai100': 0,
                'stock_diesel_winter': 0, 'stock_diesel_arctic': 0, 'stock_diesel_summer': 0,
                'transit_ai92': 0, 'transit_ai95': 0, 'transit_ai98_ai100': 0,
                'transit_diesel_winter': 0, 'transit_diesel_arctic': 0, 'transit_diesel_summer': 0,
                'capacity_ai92': 0, 'capacity_ai95': 0, 'capacity_ai98_ai100': 0,
                'capacity_diesel_winter': 0, 'capacity_diesel_arctic': 0, 'capacity_diesel_summer': 0,
            }
            non_azs_locations = []

            for loc in sheet3_recs:
                loc_name = str(loc.get('location_name', '')).lower()
                if 'азс' in loc_name:
                    azs_count += 1
                    for key in azs_totals.keys():
                        azs_totals[key] += float(loc.get(key, 0) or 0)
                else:
                    non_azs_locations.append(loc)

            # 1. СНАЧАЛА записываем объекты (нефтебазы)
            for loc in non_azs_locations:
                self._set_cell_value(ws, current_row, 1, company_name, start_row)
                # Для нефтебаз колонку B заполняем
                self._set_cell_value(ws, current_row, 2, self._get_supplier_string(company_name), start_row)
                self._set_cell_value(ws, current_row, 3, loc.get('location_name', ''), start_row)
                
                self._set_cell_value(ws, current_row, 5, loc.get('stock_ai92', 0), start_row)
                self._set_cell_value(ws, current_row, 6, loc.get('stock_ai95', 0), start_row)
                self._set_cell_value(ws, current_row, 7, loc.get('stock_ai98_ai100', 0), start_row)
                self._set_cell_value(ws, current_row, 8, loc.get('stock_diesel_winter', 0), start_row)
                self._set_cell_value(ws, current_row, 9, loc.get('stock_diesel_arctic', 0), start_row)
                self._set_cell_value(ws, current_row, 10, loc.get('stock_diesel_summer', 0), start_row)
                
                self._set_cell_value(ws, current_row, 13, loc.get('transit_ai92', 0), start_row)
                self._set_cell_value(ws, current_row, 14, loc.get('transit_ai95', 0), start_row)
                self._set_cell_value(ws, current_row, 15, loc.get('transit_ai98_ai100', 0), start_row)
                self._set_cell_value(ws, current_row, 16, loc.get('transit_diesel_winter', 0), start_row)
                self._set_cell_value(ws, current_row, 17, loc.get('transit_diesel_arctic', 0), start_row)
                self._set_cell_value(ws, current_row, 18, loc.get('transit_diesel_summer', 0), start_row)
                
                self._set_cell_value(ws, current_row, 21, loc.get('capacity_ai92', 0), start_row)
                self._set_cell_value(ws, current_row, 22, loc.get('capacity_ai95', 0), start_row)
                self._set_cell_value(ws, current_row, 23, loc.get('capacity_ai98_ai100', 0), start_row)
                self._set_cell_value(ws, current_row, 24, loc.get('capacity_diesel_winter', 0), start_row)
                self._set_cell_value(ws, current_row, 25, loc.get('capacity_diesel_arctic', 0), start_row)
                self._set_cell_value(ws, current_row, 26, loc.get('capacity_diesel_summer', 0), start_row)
                
                current_row += 1

            # 2. ЗАТЕМ записываем сведенную строку АЗС (под нефтебазами)
            if azs_count > 0:
                self._set_cell_value(ws, current_row, 1, company_name, start_row)
                # Колонка 2 (B) для строки с АЗС ПРОПУСКАЕТСЯ по просьбе пользователя
                self._set_cell_value(ws, current_row, 3, f"АЗС ({azs_count} шт)", start_row)
                
                self._set_cell_value(ws, current_row, 5, azs_totals['stock_ai92'], start_row)
                self._set_cell_value(ws, current_row, 6, azs_totals['stock_ai95'], start_row)
                self._set_cell_value(ws, current_row, 7, azs_totals['stock_ai98_ai100'], start_row)
                self._set_cell_value(ws, current_row, 8, azs_totals['stock_diesel_winter'], start_row)
                self._set_cell_value(ws, current_row, 9, azs_totals['stock_diesel_arctic'], start_row)
                self._set_cell_value(ws, current_row, 10, azs_totals['stock_diesel_summer'], start_row)
                
                self._set_cell_value(ws, current_row, 13, azs_totals['transit_ai92'], start_row)
                self._set_cell_value(ws, current_row, 14, azs_totals['transit_ai95'], start_row)
                self._set_cell_value(ws, current_row, 15, azs_totals['transit_ai98_ai100'], start_row)
                self._set_cell_value(ws, current_row, 16, azs_totals['transit_diesel_winter'], start_row)
                self._set_cell_value(ws, current_row, 17, azs_totals['transit_diesel_arctic'], start_row)
                self._set_cell_value(ws, current_row, 18, azs_totals['transit_diesel_summer'], start_row)
                
                self._set_cell_value(ws, current_row, 21, azs_totals['capacity_ai92'], start_row)
                self._set_cell_value(ws, current_row, 22, azs_totals['capacity_ai95'], start_row)
                self._set_cell_value(ws, current_row, 23, azs_totals['capacity_ai98_ai100'], start_row)
                self._set_cell_value(ws, current_row, 24, azs_totals['capacity_diesel_winter'], start_row)
                self._set_cell_value(ws, current_row, 25, azs_totals['capacity_diesel_arctic'], start_row)
                self._set_cell_value(ws, current_row, 26, azs_totals['capacity_diesel_summer'], start_row)
                
                current_row += 1

    def _fill_supply_sheet_full(self, ws, aggregated_data: dict, report_date: date):
        start_row = 9
        current_row = start_row
        current_month = report_date.strftime('%m')
        current_year = report_date.strftime('%Y')
        static_date = f"28.{current_month}.{current_year}"
        
        for company_name, company_data in aggregated_data.items():
            sheet4_recs = company_data.get('sheet4_data', [])
            if not sheet4_recs:
                continue
            
            totals = {
                'supply_ai92': 0, 'supply_ai95': 0, 'supply_ai98_100': 0,
                'supply_diesel_winter': 0, 'supply_diesel_arctic': 0, 'supply_diesel_summer': 0
            }
            
            for supply in sheet4_recs:
                for key in totals.keys():
                    totals[key] += float(supply.get(key, 0) or 0)
            
            # Колонка 1: Поставщик (из нашей функции)
            self._set_cell_value(ws, current_row, 1, self._get_supplier_string(company_name), start_row)
            # Колонка 2: Название компании
            self._set_cell_value(ws, current_row, 2, company_name, start_row)
            # Колонка 3: Нефтебаза - берем из функции _get_oil_depot_string
            self._set_cell_value(ws, current_row, 3, self._get_oil_depot_string(company_name), start_row)
            # Колонка 4: Дата поставки
            self._set_cell_value(ws, current_row, 4, static_date, start_row)
            
            self._set_cell_value(ws, current_row, 6, totals['supply_ai92'], start_row)
            self._set_cell_value(ws, current_row, 7, totals['supply_ai95'], start_row)
            self._set_cell_value(ws, current_row, 8, totals['supply_ai98_100'], start_row)
            self._set_cell_value(ws, current_row, 9, totals['supply_diesel_winter'], start_row)
            self._set_cell_value(ws, current_row, 10, totals['supply_diesel_arctic'], start_row)
            self._set_cell_value(ws, current_row, 11, totals['supply_diesel_summer'], start_row)
            
            current_row += 1

    def _fill_sales_sheet_full(self, ws, aggregated_data: dict):
        start_row = 9
        current_row = start_row
        for company_name, company_data in aggregated_data.items():
            sheet5_recs = company_data.get('sheet5_data', [])
            
            azs_count = 0
            azs_totals = {
                'daily_ai92': 0, 'daily_ai95': 0, 'daily_ai98_100': 0,
                'daily_winter': 0, 'daily_arctic': 0, 'daily_summer': 0,
                'monthly_ai92': 0, 'monthly_ai95': 0, 'monthly_ai98_100': 0,
                'monthly_diesel_winter': 0, 'monthly_diesel_arctic': 0, 'monthly_diesel_summer': 0
            }
            non_azs_locations = []
            
            for loc in sheet5_recs:
                loc_name = str(loc.get('location_name', '')).lower()
                if 'азс' in loc_name:
                    azs_count += 1
                    for key in azs_totals.keys():
                        azs_totals[key] += float(loc.get(key, 0) or 0)
                else:
                    non_azs_locations.append(loc)
                    
            # 1. СНАЧАЛА записываем остальные объекты (нефтебазы)
            for loc in non_azs_locations:
                self._set_cell_value(ws, current_row, 1, company_name, start_row)
                # Для нефтебаз колонку B заполняем
                self._set_cell_value(ws, current_row, 2, self._get_supplier_string(company_name), start_row)
                self._set_cell_value(ws, current_row, 3, loc.get('location_name', ''), start_row)
                
                self._set_cell_value(ws, current_row, 5, loc.get('daily_ai92', 0), start_row)
                self._set_cell_value(ws, current_row, 6, loc.get('daily_ai95', 0), start_row)
                self._set_cell_value(ws, current_row, 7, loc.get('daily_ai98_100', 0), start_row)
                self._set_cell_value(ws, current_row, 8, loc.get('daily_winter', 0), start_row)
                self._set_cell_value(ws, current_row, 9, loc.get('daily_arctic', 0), start_row)
                self._set_cell_value(ws, current_row, 10, loc.get('daily_summer', 0), start_row)
                
                self._set_cell_value(ws, current_row, 13, loc.get('monthly_ai92', 0), start_row)
                self._set_cell_value(ws, current_row, 14, loc.get('monthly_ai95', 0), start_row)
                self._set_cell_value(ws, current_row, 15, loc.get('monthly_ai98_100', 0), start_row)
                self._set_cell_value(ws, current_row, 16, loc.get('monthly_diesel_winter', 0), start_row)
                self._set_cell_value(ws, current_row, 17, loc.get('monthly_diesel_arctic', 0), start_row)
                self._set_cell_value(ws, current_row, 18, loc.get('monthly_diesel_summer', 0), start_row)
                
                current_row += 1

            # 2. ЗАТЕМ записываем сведенную строку АЗС (под нефтебазами)
            if azs_count > 0:
                self._set_cell_value(ws, current_row, 1, company_name, start_row)
                # Колонка 2 (B) для строки с АЗС ПРОПУСКАЕТСЯ по просьбе пользователя
                self._set_cell_value(ws, current_row, 3, f"АЗС ({azs_count} шт)", start_row)
                
                self._set_cell_value(ws, current_row, 5, azs_totals['daily_ai92'], start_row)
                self._set_cell_value(ws, current_row, 6, azs_totals['daily_ai95'], start_row)
                self._set_cell_value(ws, current_row, 7, azs_totals['daily_ai98_100'], start_row)
                self._set_cell_value(ws, current_row, 8, azs_totals['daily_winter'], start_row)
                self._set_cell_value(ws, current_row, 9, azs_totals['daily_arctic'], start_row)
                self._set_cell_value(ws, current_row, 10, azs_totals['daily_summer'], start_row)
                
                self._set_cell_value(ws, current_row, 13, azs_totals['monthly_ai92'], start_row)
                self._set_cell_value(ws, current_row, 14, azs_totals['monthly_ai95'], start_row)
                self._set_cell_value(ws, current_row, 15, azs_totals['monthly_ai98_100'], start_row)
                self._set_cell_value(ws, current_row, 16, azs_totals['monthly_diesel_winter'], start_row)
                self._set_cell_value(ws, current_row, 17, azs_totals['monthly_diesel_arctic'], start_row)
                self._set_cell_value(ws, current_row, 18, azs_totals['monthly_diesel_summer'], start_row)
                
                current_row += 1

    def _fill_aviation_sheet_full(self, ws, aggregated_data: dict):
        start_row = 8
        current_row = start_row
        for company_name, company_data in aggregated_data.items():
            for item in company_data.get('sheet6_data', []):
                self._set_cell_value(ws, current_row, 1, item.get('airport_name', ''), start_row)
                self._set_cell_value(ws, current_row, 2, item.get('tzk_name', ''), start_row)
                self._set_cell_value(ws, current_row, 3, item.get('contracts_info', ''), start_row)
                self._set_cell_value(ws, current_row, 4, item.get('supply_week', 0), start_row)
                self._set_cell_value(ws, current_row, 5, item.get('supply_month_start', 0), start_row)
                self._set_cell_value(ws, current_row, 6, item.get('monthly_demand', 0), start_row)
                self._set_cell_value(ws, current_row, 7, item.get('consumption_week', 0), start_row)
                self._set_cell_value(ws, current_row, 8, item.get('consumption_month_start', 0), start_row)
                self._set_cell_value(ws, current_row, 9, item.get('end_of_day_balance', 0), start_row)
                current_row += 1

def generate_complete_report(db_connection, template_path=None):
    generator = TemplateReportGenerator(db_connection, template_path)
    return generator.generate_report()

if __name__ == "__main__":
    from database.queries import db
    report_path = generate_complete_report(db)
    print(f"Отчет готов: {report_path}")
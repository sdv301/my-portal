# parser/unified_parser.py
import openpyxl
from openpyxl.utils import range_boundaries
from datetime import datetime
from typing import Dict, List, Any
import os
import re

class UnifiedParser:
    """Улучшенный парсер для сложных Excel файлов"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = None
        self.merged_cell_ranges = {}
    
    # В методе parse_all() unified_parser.py
    def parse_all(self) -> Dict[str, Any]:
        """Парсинг всех листов файла"""
        try:
            print(f"🧠 Начинаем парсинг файла: {self.file_path}")
            
            if not os.path.exists(self.file_path):
                raise FileNotFoundError(f"Файл не найден: {self.file_path}")
            
            # Загружаем файл
            self.wb = openpyxl.load_workbook(self.file_path, data_only=False)
            
            # Кэшируем объединенные ячейки
            self._cache_merged_cells()
            
            result = {
                'metadata': self._parse_metadata(),
                # 'sheet1': self._parse_sheet1(),  # Структура - список
                # 'sheet2': self._parse_sheet2(),  # Потребность - словарь
                'sheet3': self._parse_sheet3(),  # Остатки - список
                'sheet4': self._parse_sheet4(),  # Поставки - список
                'sheet5': self._parse_sheet5(),  # Реализация - список
                'sheet6': self._parse_sheet6(),  # Авиатопливо - список
                # 'sheet7': self._parse_sheet7(),  # Справка - список
            }
            
            print("✅ Парсинг завершен успешно!")
            return result
            
        except Exception as e:
            print(f"❌ Ошибка парсинга: {e}")
            import traceback
            traceback.print_exc()
            return self._fallback_parse()
    
    # В unified_parser.py - улучшаем метод _parse_metadata и _detect_company_from_content

    def _parse_metadata(self) -> Dict[str, Any]:
        """Улучшенное определение компании по имени файла и содержимому"""
        filename = os.path.basename(self.file_path).lower()
        
        # Более точное определение по имени файла с приоритетом
        company_mappings = [
            # Саханефтегазсбыт - высший приоритет
            ('саханефтегазсбыт', 'Саханефтегазсбыт'),
            ('снгс', 'Саханефтегазсбыт'),
            ('санги', 'Саханефтегазсбыт'),
            ('sngs', 'Саханефтегазсбыт'),
            
            # Туймаада-Нефть
            ('туймаада', 'Туймаада-Нефть'),
            ('туймааданефть', 'Туймаада-Нефть'),
            ('tumaada', 'Туймаада-Нефть'),
            
            # Сибойл
            ('сибойл', 'Сибойл'),
            ('сибирьойл', 'Сибойл'),
            ('сибирь ойл', 'Сибойл'),
            ('siboil', 'Сибойл'),
            
            # ЭКТО-Ойл
            ('экто-ойл', 'ЭКТО-Ойл'),
            ('эктоойл', 'ЭКТО-Ойл'),
            ('экто', 'ЭКТО-Ойл'),
            ('ecto-oil', 'ЭКТО-Ойл'),
            
            # Сибирское топливо
            ('сибирское', 'Сибирское топливо'),
            ('сибтопливо', 'Сибирское топливо'),
            ('sibtoplivo', 'Сибирское топливо'),
            
            # Паритет
            ('паритет', 'Паритет'),
            ('paritet', 'Паритет'),
            
        ]
        
        # Сначала проверяем имя файла
        company = 'Неизвестная компания'
        for pattern, comp_name in company_mappings:
            if pattern in filename:
                company = comp_name
                print(f"🔍 Компания определена по имени файла: {comp_name}")
                break
        
        # Если не нашли по имени файла, проверяем содержимое
        if company == 'Неизвестная компания':
            company_from_content = self._detect_company_from_content()
            if company_from_content != 'Неизвестная компания':
                company = company_from_content
                print(f"🔍 Компания определена по содержимому: {company}")
        
        # Дополнительная проверка: если в названии файла есть цифры (версии, даты), но есть ключевые слова
        if company == 'Неизвестная компания':
            # Ищем комбинации ключевых слов в имени файла
            filename_words = filename.replace('_', ' ').replace('-', ' ').split()
            
            word_combinations = [
                ('саха', 'нефтегазсбыт'),
                ('сиб', 'ойл'),
                ('туймаада', 'нефть'),
                ('сибирское', 'топливо'),
            ]
            
            for word1, word2 in word_combinations:
                if word1 in filename_words and word2 in filename_words:
                    # Пытаемся определить по комбинации
                    if word1 == 'саха' and word2 == 'нефтегазсбыт':
                        company = 'Саханефтегазсбыт'
                    elif word1 == 'сиб' and word2 == 'ойл':
                        company = 'Сибойл'
                    elif word1 == 'туймаада' and word2 == 'нефть':
                        company = 'Туймаада-Нефть'
                    elif word1 == 'сибирское' and word2 == 'топливо':
                        company = 'Сибирское топливо'
                    
                    if company != 'Неизвестная компания':
                        print(f"🔍 Компания определена по комбинации слов: {company}")
                        break
        
        return {
            'company': company,
            'report_date': datetime.now(),
            'filename': filename,
            'sheets_available': self.wb.sheetnames if self.wb else []
        }

    def _detect_company_from_content(self) -> str:
        """Улучшенное определение компании по содержимому файла"""
        try:
            # Проверяем все листы, а не только первый
            for sheet_name in self.wb.sheetnames:
                ws = self.wb[sheet_name]
                
                # Ищем в первых 50 строках каждого листа
                for row in range(1, min(51, ws.max_row + 1)):
                    for col in range(1, min(10, ws.max_column + 1)):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value and isinstance(cell_value, str):
                            cell_value_lower = cell_value.lower()
                            
                            # Расширенный список ключевых слов для поиска
                            if any(name in cell_value_lower for name in [
                                'саханефтегазсбыт', 'снгс', 'ао "саханефтегазсбыт"',
                                'санги', 'саха нефтегазсбыт'
                            ]):
                                return 'Саханефтегазсбыт'
                            elif any(name in cell_value_lower for name in [
                                'туймаада-нефть', 'туймаада нефть', 'ао нк "туймаада-нефть"'
                            ]):
                                return 'Туймаада-Нефть'
                            elif any(name in cell_value_lower for name in [
                                'сибойл', 'сибирьойл', 'ооо "сибирьойл"', 'сибирь ойл'
                            ]):
                                return 'Сибойл'
                            elif any(name in cell_value_lower for name in [
                                'экто-ойл', 'эктоойл', 'ооо "экто-ойл"', 'экто ойл'
                            ]):
                                return 'ЭКТО-Ойл'
                            elif any(name in cell_value_lower for name in [
                                'сибирское топливо', 'сибтопливо'
                            ]):
                                return 'Сибирское топливо'
                            elif any(name in cell_value_lower for name in [
                                'паритет', 'ооо "паритет"'
                            ]):
                                return 'Паритет'
                            
            return 'Неизвестная компания'
        except Exception as e:
            print(f"⚠️ Ошибка при определении компании из содержимого: {e}")
            return 'Неизвестная компания'
    
    # def _parse_sheet1(self) -> List[Dict[str, Any]]:
    #     """Парсинг Листа 1: Структура"""
    #     try:
    #         ws = self.wb['1-Структура']
    #         data = []
            
    #         print("🔍 Парсим Лист 1 (Структура)...")
            
    #         # Данные компаний: строки 11-19
    #         for row_num in range(11, 20):
    #             row_data = {}
                
    #             # Колонка A: Принадлежность
    #             cell_a = ws.cell(row=row_num, column=1)
    #             affiliation = self._safe_str(cell_a.value)
    #             if not affiliation or affiliation == '':
    #                 continue
                
    #             row_data['affiliation'] = affiliation
                
    #             # Колонка B: Компания-поставщик
    #             cell_b = ws.cell(row=row_num, column=2)
    #             row_data['company'] = self._safe_str(cell_b.value)
                
    #             # Колонки C, D, E: числовые данные
    #             row_data['oil_depots_count'] = self._safe_int(ws.cell(row=row_num, column=3).value)
    #             row_data['azs_count'] = self._safe_int(ws.cell(row=row_num, column=4).value)
    #             row_data['working_azs_count'] = self._safe_int(ws.cell(row=row_num, column=5).value)
                
    #             if row_data['company']:
    #                 data.append(row_data)
            
    #         print(f"✅ Лист 1 обработан: {len(data)} записей")
    #         return data
            
    #     except Exception as e:
    #         print(f"❌ Ошибка парсинга Листа 1: {e}")
    #         return []
    
    # def _parse_sheet2(self) -> Dict[str, Any]:
    #     """Улучшенный парсинг Листа 2: Потребность"""
    #     try:
    #         ws = self.wb['2-Потребность']
    #         data = {}
            
    #         print("🔍 Парсим Лист 2 (Потребность)...")
            
    #         # Ищем строки с "ГОД" и "МЕСЯЦ"
    #         for row_num in range(1, 20):
    #             cell_value = ws.cell(row=row_num, column=1).value
    #             if cell_value and isinstance(cell_value, str):
    #                 cell_str = str(cell_value).upper()
                    
    #                 if 'ГОД' in cell_str:
    #                     # Годовая потребность - берем данные из этой строки
    #                     row = ws[row_num]
    #                     if len(row) >= 11:
    #                         data.update({
    #                             'yearly_gasoline_total': self._safe_float(row[1].value),
    #                             'yearly_ai92': self._safe_float(row[3].value),
    #                             'yearly_ai95': self._safe_float(row[4].value),
    #                             'yearly_ai98_100': self._safe_float(row[5]),
    #                             'yearly_diesel_total': self._safe_float(row[6].value),
    #                             'yearly_diesel_winter': self._safe_float(row[7]),
    #                             'yearly_diesel_arctic': self._safe_float(row[8]),
    #                             'yearly_diesel_summer': self._safe_float(row[9]),
    #                             'yearly_diesel_intermediate': self._safe_float(row[10]),
    #                         })
                    
    #                 elif 'МЕСЯЦ' in cell_str:
    #                     # Месячная потребность - берем данные из этой строки
    #                     row = ws[row_num]
    #                     if len(row) >= 11:
    #                         data.update({
    #                             'monthly_gasoline_total': self._safe_float(row[1].value),
    #                             'monthly_ai92': self._safe_float(row[3].value),
    #                             'monthly_ai95': self._safe_float(row[4].value),
    #                             'monthly_ai98_100': self._safe_float(row[5]),
    #                             'monthly_diesel_total': self._safe_float(row[6].value),
    #                             'monthly_diesel_winter': self._safe_float(row[7]),
    #                             'monthly_diesel_arctic': self._safe_float(row[8]),
    #                             'monthly_diesel_summer': self._safe_float(row[9]),
    #                             'monthly_diesel_intermediate': self._safe_float(row[10])
    #                         })
            
    #         print(f"✅ Лист 2 обработан: {len(data)} показателей")
    #         return data
            
    #     except Exception as e:
    #         print(f"❌ Ошибка парсинга Листа 2: {e}")
    #         return {}

    def _parse_sheet3(self) -> List[Dict[str, Any]]:
        """Исправленный парсинг Листа 3: Остатки с учетом реальной структуры"""
        try:
            ws = self.wb['3-Остатки']
            data = []
            
            print("🔍 Парсим Лист 3 (Остатки) с учетом реальной структуры...")
            
            # Реальная структура: данные начинаются со строки 9
            # Колонка B: "ВИНК" (группировка)
            # Колонка C: Название компании  
            # Колонки D-K: Числовые данные
            
            for row_num in range(9, ws.max_row + 1):
                row_data = {}
                
                # Колонка C: Название компании
                cell_c = ws.cell(row=row_num, column=3)  # Колонка C
                company = self._safe_str(cell_c.value)
                
                if not company or company == '' or company in ['1', '2', '3']:
                    continue
                    
                row_data['company'] = company
                
                # Колонка B: Группировка ("ВИНК")
                cell_b = ws.cell(row=row_num, column=2)  # Колонка B
                row_data['group'] = self._safe_str(cell_b.value)
                
                # Колонка D: Объект (из контекста - это нефтебаза/АЗС)
                cell_d = ws.cell(row=row_num, column=4)  # Колонка D
                row_data['object_name'] = self._safe_str(cell_d.value)
                
                # Числовые данные (колонки E-L соответствуют D-K в вашем описании)
                
                row_data['stock_ai92'] = self._get_cell_value(ws, row_num, 5)      # F
                row_data['stock_ai95'] = self._get_cell_value(ws, row_num, 6)      # G  
                row_data['stock_ai98_100'] = self._get_cell_value(ws, row_num, 7)      # G  
                row_data['stock_diesel_winter'] = self._get_cell_value(ws, row_num, 8)   # I
                row_data['stock_diesel_arctic'] = self._get_cell_value(ws, row_num, 9)  # J
                row_data['stock_diesel_summer'] = self._get_cell_value(ws, row_num, 10)  # K
                
                row_data['transit_ai92'] = self._get_cell_value(ws, row_num, 13)      # F
                row_data['transit_ai95'] = self._get_cell_value(ws, row_num, 14)      # G  
                row_data['transit_ai98_100'] = self._get_cell_value(ws, row_num, 15)      # G  
                row_data['transit_diesel_winter'] = self._get_cell_value(ws, row_num, 16)   # I
                row_data['transit_diesel_arctic'] = self._get_cell_value(ws, row_num, 17)  # J
                row_data['transit_diesel_summer'] = self._get_cell_value(ws, row_num, 19)  # K
                
                row_data['capacity_ai92'] = self._get_cell_value(ws, row_num, 21)      # F
                row_data['capacity_ai95'] = self._get_cell_value(ws, row_num, 22)      # G  
                row_data['capacity_ai98_100'] = self._get_cell_value(ws, row_num, 23)      # G  
                row_data['capacity_diesel_winter'] = self._get_cell_value(ws, row_num, 24)   # I
                row_data['capacity_diesel_arctic'] = self._get_cell_value(ws, row_num, 25)  # J
                row_data['capacity_diesel_summer'] = self._get_cell_value(ws, row_num, 26)  # K
                
                # Добавляем только если есть значимые данные
                significant_keys = [
                    'stock_ai92', 'stock_ai95', 'stock_ai98_100', 'stock_diesel_winter', 'stock_diesel_arctic', 'stock_diesel_summer',
                    'transit_ai92', 'transit_ai95', 'transit_ai98_100', 'transit_diesel_winter', 'transit_diesel_arctic', 'transit_diesel_summer'
                ]
                if (row_data['company'] and 
                    any(row_data.get(key, 0) > 0 for key in significant_keys)):
                    data.append(row_data)
                    print(f"   📊 Найдены данные: {row_data['company']} - АИ-92: {row_data.get('stock_ai92', 0)}, АИ-95: {row_data.get('stock_ai95', 0)}")
            
            print(f"✅ Лист 3 обработан: {len(data)} записей")
            return data
            
        except Exception as e:
            print(f"❌ Ошибка парсинга Листа 3: {e}")
            import traceback
            traceback.print_exc()
            return []

    def _parse_sheet4(self) -> List[Dict[str, Any]]:
        """Парсинг Листа 4: Поставки"""
        try:
            ws = self.wb['4-Поставка']
            data = []
            current_company = None
            
            print("🔍 Парсим Лист 4 (Поставки)...")
            
            for row_num in range(6, ws.max_row + 1):
                row_data = {}
                
                # Обработка компании (аналогично листу 3)
                merged_value = self._get_merged_cell_value('4-Поставка', row_num, 1)
                if merged_value and merged_value != '':
                    current_company = str(merged_value).strip()
                    row_data['company'] = current_company
                else:
                    cell_a = ws.cell(row=row_num, column=1)
                    if cell_a.value and str(cell_a.value).strip() != '':
                        current_company = str(cell_a.value).strip()
                        row_data['company'] = current_company
                    elif current_company:
                        row_data['company'] = current_company
                    else:
                        continue
                
                # Фильтруем некорректные названия
                if current_company in ['1', '2', '3', '4', '5', '6', '7', '8', '9']:
                    continue
                
                # Колонка B: Компания (дублирующая информация)
                cell_b = ws.cell(row=row_num, column=2)
                row_data['company_duplicate'] = self._safe_str(cell_b.value)
                
                # Колонка C: Нефтебаза
                cell_c = ws.cell(row=row_num, column=3)
                row_data['oil_depot'] = self._safe_str(cell_c.value)
                
                # Колонка D: Срок поставки
                cell_d = ws.cell(row=row_num, column=4)
                row_data['supply_date'] = self._safe_str(cell_d.value)
                
                # Числовые данные поставок
                
                row_data['supply_ai92'] = self._get_cell_value(ws, row_num, 6)
                row_data['supply_ai95'] = self._get_cell_value(ws, row_num, 7)
                row_data['supply_ai98_100'] = self._get_cell_value(ws, row_num, 8)

                row_data['supply_diesel_winter'] = self._get_cell_value(ws, row_num, 9)
                row_data['supply_diesel_arctic'] = self._get_cell_value(ws, row_num, 10)
                row_data['supply_diesel_summer'] = self._get_cell_value(ws, row_num, 11)
                
                if row_data['company']:
                    data.append(row_data)
            
            print(f"✅ Лист 4 обработан: {len(data)} записей")
            return data
            
        except Exception as e:
            print(f"❌ Ошибка парсинга Листа 4: {e}")
            return []
    
    def _parse_sheet5(self) -> List[Dict[str, Any]]:
        """Парсинг Листа 5: Реализация"""
        try:
            ws = self.wb['5-Реализация']
            data = []
            current_company = None
            
            print("🔍 Парсим Лист 5 (Реализация)...")
            
            for row_num in range(9, ws.max_row + 1):
                row_data = {}
                
                # Обработка компании
                merged_value = self._get_merged_cell_value('5-Реализация', row_num, 1)
                if merged_value and merged_value != '':
                    current_company = str(merged_value).strip()
                    row_data['company'] = current_company
                else:
                    cell_a = ws.cell(row=row_num, column=1)
                    if cell_a.value and str(cell_a.value).strip() != '':
                        current_company = str(cell_a.value).strip()
                        row_data['company'] = current_company
                    elif current_company:
                        row_data['company'] = current_company
                    else:
                        continue
                
                # Фильтруем некорректные названия
                if current_company in ['1', '2', '3', '4', '5', '6', '7', '8', '9']:
                    continue
                
                # Колонка B: Поставщик
                cell_b = ws.cell(row=row_num, column=2)
                row_data['supplier'] = self._safe_str(cell_b.value)
                
                # Колонка C: Объект
                cell_c = ws.cell(row=row_num, column=3)
                row_data['object_name'] = self._safe_str(cell_c.value)
                
                # Реализация с начала месяца (важные данные!)
                row_data['daily_ai92'] = self._get_cell_value(ws, row_num, 5)  # M
                row_data['daily_ai95'] = self._get_cell_value(ws, row_num, 6)  # N
                row_data['daily_ai98_100'] = self._get_cell_value(ws, row_num, 7)  # N
                row_data['daily_winter'] = self._get_cell_value(ws, row_num, 8)
                row_data['daily_arctic'] = self._get_cell_value(ws, row_num, 9)
                row_data['daily_summer'] = self._get_cell_value(ws, row_num, 10)
                
                row_data['monthly_ai92'] = self._get_cell_value(ws, row_num, 13)  # M
                row_data['monthly_ai95'] = self._get_cell_value(ws, row_num, 14)  # N
                row_data['monthly_ai98_100'] = self._get_cell_value(ws, row_num, 15)  # N
                row_data['monthly_winter'] = self._get_cell_value(ws, row_num, 16)
                row_data['monthly_arctic'] = self._get_cell_value(ws, row_num, 17)
                row_data['monthly_summer'] = self._get_cell_value(ws, row_num, 18)
                
                if row_data['company'] and (row_data['monthly_ai92'] > 0 or row_data['monthly_ai95'] > 0):
                    data.append(row_data)
            
            print(f"✅ Лист 5 обработан: {len(data)} записей")
            return data
            
        except Exception as e:
            print(f"❌ Ошибка парсинга Листа 5: {e}")
            return []
    
    def _parse_sheet6(self) -> List[Dict[str, Any]]:
        """Парсинг Листа 6: Авиатопливо"""
        try:
            ws = self.wb['6-Авиатопливо']
            data = []
            
            print("🔍 Парсим Лист 6 (Авиатопливо)...")
            
            for row_num in range(8, ws.max_row + 1):
                row_data = {}
                
                # Колонка A: Аэропорт
                cell_a = ws.cell(row=row_num, column=1)
                airport = self._safe_str(cell_a.value)
                if not airport:
                    continue
                
                row_data['airport'] = airport
                
                # Колонка B: ТЗК
                cell_b = ws.cell(row=row_num, column=2)
                row_data['tzk'] = self._safe_str(cell_b.value)
                
                # Колонка C: Договоры
                cell_c = ws.cell(row=row_num, column=3)
                row_data['contracts'] = self._safe_str(cell_c.value)
                
                # Числовые данные
                row_data['supply_week'] = self._get_cell_value(ws, row_num, 4)
                row_data['supply_month_start'] = self._get_cell_value(ws, row_num, 5)
                row_data['monthly_demand'] = self._get_cell_value(ws, row_num, 6)
                row_data['consumption_week'] = self._get_cell_value(ws, row_num, 7)
                row_data['consumption_month_start'] = self._get_cell_value(ws, row_num, 8)
                row_data['end_of_day_balance'] = self._get_cell_value(ws, row_num, 9)
                
                data.append(row_data)
            
            print(f"✅ Лист 6 обработан: {len(data)} записей")
            return data
            
        except Exception as e:
            print(f"❌ Ошибка парсинга Листа 6: {e}")
            return []
    
    def _parse_sheet7(self) -> List[Dict[str, Any]]:
        """Парсинг Листа 7: Справка"""
        try:
            ws = self.wb['7-Справка']
            data = []
            
            print("🔍 Парсим Лист 7 (Справка)...")
            
            for row_num in range(6, ws.max_row + 1):
                row_data = {}
                
                # Колонка A: Топливо
                cell_a = ws.cell(row=row_num, column=1)
                fuel_type = self._safe_str(cell_a.value)
                if not fuel_type:
                    continue
                
                row_data['fuel_type'] = fuel_type
                
                # Колонка B: Ситуация
                cell_b = ws.cell(row=row_num, column=2)
                row_data['situation'] = self._safe_str(cell_b.value)
                
                # Колонка C: Комментарии
                cell_c = ws.cell(row=row_num, column=3)
                row_data['comments'] = self._safe_str(cell_c.value)
                
                data.append(row_data)
            
            print(f"✅ Лист 7 обработан: {len(data)} записей")
            return data
            
        except Exception as e:
            print(f"❌ Ошибка парсинга Листа 7: {e}")
            return []
    
    def _cache_merged_cells(self):
        """Кэшируем информацию об объединенных ячейках"""
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            self.merged_cell_ranges[sheet_name] = {}
            
            for merged_range in ws.merged_cells.ranges:
                min_row, min_col, max_row, max_col = range_boundaries(merged_range.coord)
                first_cell = ws.cell(min_row, min_col)
                
                self.merged_cell_ranges[sheet_name][merged_range.coord] = {
                    'min_row': min_row, 'max_row': max_row,
                    'min_col': min_col, 'max_col': max_col,
                    'value': first_cell.value
                }
    
    def _get_merged_cell_value(self, sheet_name: str, row: int, col: int):
        """Получаем значение объединенной ячейки"""
        if sheet_name not in self.merged_cell_ranges:
            return None
            
        for range_info in self.merged_cell_ranges[sheet_name].values():
            if (range_info['min_row'] <= row <= range_info['max_row'] and
                range_info['min_col'] <= col <= range_info['max_col']):
                return range_info['value']
        return None
    
    def _get_cell_value(self, ws, row: int, col: int):
        """Безопасное получение значения ячейки"""
        try:
            cell = ws.cell(row=row, column=col)
            
            if cell.data_type == 'f':  # Формула
                try:
                    wb_calculated = openpyxl.load_workbook(self.file_path, data_only=True)
                    ws_calculated = wb_calculated[ws.title]
                    calculated_value = ws_calculated.cell(row=row, column=col).value
                    wb_calculated.close()
                    return self._safe_float(calculated_value)
                except:
                    return 0.0
            else:
                return self._safe_float(cell.value)
        except:
            return 0.0
    
    def _safe_str(self, value) -> str:
        """Безопасное преобразование в строку"""
        if value is None:
            return ''
        return str(value).strip()
    
    def _safe_int(self, value) -> int:
        """Безопасное преобразование в int"""
        try:
            if value is None:
                return 0
            return int(float(str(value).replace(',', '.')))
        except:
            return 0
    
    def _safe_float(self, value) -> float:
        """Безопасное преобразование в float"""
        try:
            if value is None:
                return 0.0
            return float(str(value).replace(',', '.'))
        except:
            return 0.0
    
    def _fallback_parse(self):
        """Резервный парсинг при ошибках"""
        return {
            'metadata': {'company': 'Неизвестная компания', 'report_date': datetime.now()},
            'sheet1': [], 'sheet2': {}, 'sheet3': [],
            'sheet4': [], 'sheet5': [], 'sheet6': [], 'sheet7': []
        }
